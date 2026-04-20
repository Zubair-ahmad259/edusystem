from django.http import HttpResponseForbidden
from django.shortcuts import render, get_object_or_404, redirect
from .models import Teacher
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.utils.text import slugify
from django.contrib.auth.decorators import login_required
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl import Workbook
from django.http import HttpResponse

# ============= TEACHER MANAGEMENT VIEWS =============

def add_teacher(request):
    if request.method == "POST":
        try:
            # Handle salary default to 0 if not provided
            salary = request.POST.get('salary', 0)
            if not salary or salary == '':
                salary = 0
            
            # Get all data from POST
            teacher = Teacher.objects.create(
                first_name=request.POST.get('first_name'),
                last_name=request.POST.get('last_name'),
                father_name=request.POST.get('father_name'),
                teacher_id=request.POST.get('teacher_id'),
                gender=request.POST.get('gender'),
                date_of_birth=request.POST.get('date_of_birth'),
                salary=salary,
                religion=request.POST.get('religion'),
                joining_date=request.POST.get('joining_date'),
                mobile_number=request.POST.get('mobile_number'),
                email=request.POST.get('email'),
                field=request.POST.get('field'),
                experience=request.POST.get('experience', 0),
                teacher_image=request.FILES.get('teacher_image'),
            )
            
            messages.success(request, "Teacher added successfully")
            return redirect('teacher_list')
            
        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")
    
    # Calculate max date of birth (18 years ago)
    from datetime import date
    from dateutil.relativedelta import relativedelta
    max_dob = date.today() - relativedelta(years=18)
    
    return render(request, "teacher/add-teacher.html", {'max_dob': max_dob})

def teacher_list(request):
    teachers = Teacher.objects.all().order_by('last_name', 'first_name')
    
    # Calculate statistics
    total_teachers = teachers.count()
    # total_salary = teachers.aggregate(total=models.Sum('salary'))['total'] or 0
    
    context = {
        "teachers": teachers,
        "total_teachers": total_teachers,
        # "total_salary": total_salary,
        # "avg_salary": total_salary // total_teachers if total_teachers > 0 else 0,
    }
    return render(request, "teacher/teachers.html", context)

def edit_teacher(request, teacher_id):
    teacher_obj = get_object_or_404(Teacher, teacher_id=teacher_id)
    
    if request.method == "POST":
        try:
            # Update fields
            teacher_obj.first_name = request.POST.get('first_name')
            teacher_obj.last_name = request.POST.get('last_name')
            teacher_obj.father_name = request.POST.get('father_name')
            teacher_obj.gender = request.POST.get('gender')
            teacher_obj.date_of_birth = request.POST.get('date_of_birth')
            
            # Handle salary - default to 0 if empty
            salary = request.POST.get('salary', 0)
            teacher_obj.salary = float(salary) if salary else 0
            
            teacher_obj.religion = request.POST.get('religion')
            teacher_obj.joining_date = request.POST.get('joining_date')
            teacher_obj.mobile_number = request.POST.get('mobile_number')
            teacher_obj.email = request.POST.get('email')
            teacher_obj.field = request.POST.get('field')
            teacher_obj.experience = request.POST.get('experience', 0)
            
            # Handle image upload
            if 'teacher_image' in request.FILES:
                teacher_obj.teacher_image = request.FILES['teacher_image']
            
            teacher_obj.save()
            messages.success(request, "Teacher updated successfully")
            return redirect('teacher_list')
            
        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")
    
    context = {"teacher": teacher_obj}
    return render(request, "teacher/edit-teacher.html", context)

def view_teacher(request, teacher_id):
    teacher_obj = get_object_or_404(Teacher, teacher_id=teacher_id)
    context = {"teacher": teacher_obj}
    return render(request, "teacher/teacher-details.html", context)

def delete_teacher(request, teacher_id):
    teacher_obj = get_object_or_404(Teacher, teacher_id=teacher_id)
    if request.method == "POST":
        teacher_obj.delete()
        messages.success(request, "Teacher deleted successfully")
        return redirect('teacher_list')
    return redirect('teacher_list')

# ============= BULK TEACHER IMPORT =============

def bulk_paste_teachers(request):
    """Bulk add teachers by pasting data"""
    
    if request.method == 'POST':
        pasted_data = request.POST.get('pasted_data', '')
        rows = pasted_data.strip().split('\n')
        
        success_count = 0
        error_count = 0
        errors = []
        
        # Get the delimiter
        delimiter = '\t' if '\t' in rows[0] else ','
        
        for line_num, line in enumerate(rows, start=1):
            if not line.strip():
                continue
            
            if delimiter == '\t':
                cols = [x.strip() for x in line.split('\t')]
            else:
                cols = [x.strip() for x in line.split(',')]
            
            if len(cols) < 12:
                errors.append(f"Row {line_num}: Only {len(cols)} columns found. Need at least 12 columns.")
                error_count += 1
                continue
            
            try:
                # Extract data (12+ columns)
                first_name = cols[0]
                last_name = cols[1]
                father_name = cols[2]
                teacher_id = cols[3]
                gender = cols[4]
                date_of_birth = cols[5]
                email = cols[6]
                mobile_number = cols[7]
                field = cols[8]
                experience = int(cols[9]) if cols[9].isdigit() else 0
                salary = float(cols[10]) if cols[10] else 0
                joining_date = cols[11]
                religion = cols[12] if len(cols) > 12 else ''
                
                # Check if teacher already exists
                if Teacher.objects.filter(teacher_id=teacher_id).exists():
                    raise Exception(f"Teacher ID '{teacher_id}' already exists")
                
                if Teacher.objects.filter(email=email).exists():
                    raise Exception(f"Email '{email}' already exists")
                
                # Create teacher
                teacher = Teacher.objects.create(
                    first_name=first_name,
                    last_name=last_name,
                    father_name=father_name,
                    teacher_id=teacher_id,
                    gender=gender,
                    date_of_birth=date_of_birth,
                    email=email,
                    mobile_number=mobile_number,
                    field=field,
                    experience=experience,
                    salary=salary,
                    joining_date=joining_date,
                    religion=religion,
                )
                success_count += 1
                
            except Exception as e:
                error_count += 1
                errors.append(f"Row {line_num}: {str(e)}")
        
        if success_count > 0:
            messages.success(request, f'✅ Successfully added {success_count} teachers!')
        if errors:
            for error in errors[:10]:
                messages.error(request, f'❌ {error}')
        
        return redirect('teacher_list')
    
    return render(request, 'teacher/bulk-paste.html')

def download_teacher_template(request):
    """Download Excel template for teacher import"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Teacher Import Template"
    
    headers = [
        'First Name*', 'Last Name*', 'Father Name*', 'Teacher ID*', 
        'Gender* (M/F)', 'Date of Birth* (YYYY-MM-DD)', 'Email*', 'Mobile Number*',
        'Field/Specialization*', 'Experience (Years)', 'Salary (PKR)', 'Joining Date*', 'Religion'
    ]
    
    # Style headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Add sample data
    sample_data = [
        ['John', 'Smith', 'Robert Smith', 'TCH001', 'M', '1985-05-15', 
         'john.smith@example.com', '+92 300 1234567', 'Computer Science', '8', '80000', '2020-01-15', 'Islamabad'],
        ['Sarah', 'Ahmed', 'Khalid Ahmed', 'TCH002', 'F', '1990-03-20',
         'sarah.ahmed@example.com', '+92 322 7654321', 'Mathematics', '5', '60000', '2021-08-01', 'Lahore'],
    ]
    
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Set column widths
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="teacher_import_template.xlsx"'
    wb.save(response)
    return response

def export_teachers_excel(request):
    """Export all teachers to Excel"""
    
    teachers = Teacher.objects.all().order_by('first_name')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Teachers List"
    
    headers = [
        'Teacher ID', 'First Name', 'Last Name', 'Father Name', 'Gender', 
        'Date of Birth', 'Email', 'Mobile Number', 'Field', 'Experience', 
        'Salary', 'Joining Date', 'Religion'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
    
    for row_num, teacher in enumerate(teachers, 2):
        ws.cell(row=row_num, column=1, value=teacher.teacher_id)
        ws.cell(row=row_num, column=2, value=teacher.first_name)
        ws.cell(row=row_num, column=3, value=teacher.last_name)
        ws.cell(row=row_num, column=4, value=teacher.father_name or '')
        ws.cell(row=row_num, column=5, value=teacher.gender)
        ws.cell(row=row_num, column=6, value=teacher.date_of_birth.strftime('%Y-%m-%d') if teacher.date_of_birth else '')
        ws.cell(row=row_num, column=7, value=teacher.email)
        ws.cell(row=row_num, column=8, value=teacher.mobile_number)
        ws.cell(row=row_num, column=9, value=teacher.field)
        ws.cell(row=row_num, column=10, value=teacher.experience)
        ws.cell(row=row_num, column=11, value=float(teacher.salary) if teacher.salary else 0)
        ws.cell(row=row_num, column=12, value=teacher.joining_date.strftime('%Y-%m-%d') if teacher.joining_date else '')
        ws.cell(row=row_num, column=13, value=teacher.religion or '')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="teachers_export.xlsx"'
    wb.save(response)
    return response