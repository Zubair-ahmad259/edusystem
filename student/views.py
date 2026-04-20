from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from .models import Batch, Semester, Section, Parent, Student, Discipline
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl import Workbook
import io

# ============= EXCEL IMPORT/EXPORT FUNCTIONS =============
def bulk_paste_students(request):
    
    if request.method == 'POST':
        pasted_data = request.POST.get('pasted_data', '')
        rows = pasted_data.strip().split('\n')
        
        success_count = 0
        error_count = 0
        errors = []
        
        # Get the delimiter (tab or comma)
        delimiter = '\t' if '\t' in rows[0] else ','
        
        for line_num, line in enumerate(rows, start=1):
            # Skip empty lines
            if not line.strip():
                continue
            
            # Split the line
            if delimiter == '\t':
                cols = [x.strip() for x in line.split('\t')]
            else:
                cols = [x.strip() for x in line.split(',')]
            
            # Check minimum columns (13 required fields)
            if len(cols) < 13:
                errors.append(f"Row {line_num}: Only {len(cols)} columns found. Need at least 13 columns.")
                error_count += 1
                continue
            
            try:
                # Extract data from columns
                first_name = cols[0]
                last_name = cols[1]
                student_id = cols[2]
                admission_number = cols[3]
                gender_value = cols[4].upper()
                dob = cols[5]
                email = cols[6]
                contact_number = cols[7] if len(cols) > 7 else ''
                batch_name = cols[8]
                semester_num = cols[9]
                section_name = cols[10]
                discipline_name = cols[11]
                address = cols[12]
                father_name = cols[13] if len(cols) > 13 else ''
                mother_name = cols[14] if len(cols) > 14 else ''
                father_contact = cols[15] if len(cols) > 15 else ''
                father_email = cols[16] if len(cols) > 16 else ''
                parent_address = cols[17] if len(cols) > 17 else ''
                
                # Validate required fields
                if not first_name:
                    raise Exception("First Name is required")
                if not last_name:
                    raise Exception("Last Name is required")
                if not student_id:
                    raise Exception("Student ID is required")
                if not admission_number:
                    raise Exception("Admission Number is required")
                if not dob:
                    raise Exception("Date of Birth is required")
                if not email:
                    raise Exception("Email is required")
                if not batch_name:
                    raise Exception("Batch Name is required")
                if not semester_num:
                    raise Exception("Semester Number is required")
                if not section_name:
                    raise Exception("Section Name is required")
                if not discipline_name:
                    raise Exception("Discipline Name is required")
                
                # Check for duplicates
                if Student.objects.filter(student_id=student_id).exists():
                    raise Exception(f"Student ID '{student_id}' already exists")
                
                if Student.objects.filter(email=email).exists():
                    raise Exception(f"Email '{email}' already exists")
                
                # Get or create Batch (using your batch naming convention)
                batch, batch_created = Batch.objects.get_or_create(name=batch_name)
                
                # Get or create Semester (1-8 only)
                try:
                    semester_num_int = int(float(semester_num))
                    if semester_num_int < 1 or semester_num_int > 8:
                        raise Exception("Semester must be between 1 and 8")
                except:
                    semester_num_int = 1
                semester, semester_created = Semester.objects.get_or_create(number=semester_num_int)
                
                # Get or create Section (must exist with batch)
                section, section_created = Section.objects.get_or_create(
                    name=section_name, 
                    batch=batch
                )
                
                # Get or create Discipline (using your exact discipline names)
                discipline, discipline_created = Discipline.objects.get_or_create(
                    field=discipline_name,
                    defaults={'program': 'BS'}  # Default program
                )
                
                # Convert gender
                if gender_value in ['M', 'MALE']:
                    gender = 'M'
                elif gender_value in ['F', 'FEMALE']:
                    gender = 'F'
                else:
                    gender = 'O'
                
                # Parse date
                try:
                    if '-' in dob:
                        dob_date = datetime.strptime(dob, '%Y-%m-%d').date()
                    elif '/' in dob:
                        dob_date = datetime.strptime(dob, '%d/%m/%Y').date()
                    else:
                        dob_date = datetime.strptime(dob, '%Y-%m-%d').date()
                except:
                    raise Exception(f"Invalid date format. Use YYYY-MM-DD")
                
                # Create Parent
                parent = None
                if father_name:
                    parent = Parent.objects.create(
                        father_name=father_name,
                        mother_name=mother_name,
                        father_contact=father_contact,
                        father_email=father_email,
                        address=parent_address or address
                    )
                
                # Create Student
                student = Student(
                    first_name=first_name,
                    last_name=last_name,
                    student_id=student_id,
                    admission_number=admission_number,
                    gender=gender,
                    dob=dob_date,
                    email=email,
                    contact_number=contact_number,
                    batch=batch,
                    semester=semester,
                    section=section,
                    discipline=discipline,
                    parent=parent,
                    address=address
                )
                student.save()
                success_count += 1
                
            except Exception as e:
                error_count += 1
                errors.append(f"Row {line_num}: {str(e)}")
        
        # Show results
        if success_count > 0:
            messages.success(request, f'✅ Successfully added {success_count} students!')
        if errors:
            for error in errors[:10]:
                messages.error(request, f'❌ {error}')
            if len(errors) > 10:
                messages.warning(request, f'And {len(errors) - 10} more errors...')
        
        return redirect('student_list')
    
    # GET request - show the paste form
    return render(request, 'students/bulk-paste.html')
def download_excel_template(request):
    """Download Excel template for bulk student import"""
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Import Template"
    
    # Define headers
    headers = [
        'First Name*', 'Last Name*', 'Student ID*', 'Admission Number*',
        'Gender* (M/F/O)', 'Date of Birth* (YYYY-MM-DD)', 'Email*', 'Contact Number',
        'Batch Name*', 'Semester Number*', 'Section Name*', 'Discipline Name*',
        'Address*', 'Father Name', 'Mother Name', 'Father Contact', 'Father Email', 'Parent Address'
    ]
    
    # Style headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Add sample data
    sample_data = [
        ['John', 'Doe', 'STU001', 'ADM001', 'M', '2000-01-15', 
         'john.doe@example.com', '1234567890', 'Batch 2024', '1', 
         'Section A', 'Computer Science', '123 Main St', 'Robert Doe', 
         'Jane Doe', '0987654321', 'robert@example.com', '123 Main St'],
        ['Jane', 'Smith', 'STU002', 'ADM002', 'F', '2001-03-20',
         'jane.smith@example.com', '0987654321', 'Batch 2024', '1',
         'Section B', 'Computer Science', '456 Oak Ave', 'Michael Smith',
         'Sarah Smith', '1234567890', 'michael@example.com', '456 Oak Ave']
    ]
    
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Set column widths
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="student_import_template.xlsx"'
    wb.save(response)
    return response


def download_student_template(request):
    """Alias for download_excel_template - for template compatibility"""
    return download_excel_template(request)


def bulk_student_import(request):
    """Handle bulk student import from Excel/CSV file"""
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        # Check file extension
        file_extension = excel_file.name.split('.')[-1].lower()
        if file_extension not in ['xlsx', 'xls', 'csv']:
            messages.error(request, 'Please upload an Excel (.xlsx, .xls) or CSV file')
            return redirect('add_student')
        
        try:
            # Read the file based on extension
            if file_extension == 'csv':
                df = pd.read_csv(excel_file)
            else:
                df = pd.read_excel(excel_file)
            
            success_count = 0
            error_count = 0
            errors = []
            warnings = []
            
            # Normalize column names (remove * and extra spaces)
            df.columns = df.columns.str.replace('*', '').str.strip()
            
            for index, row in df.iterrows():
                try:
                    # Skip empty rows
                    if pd.isna(row.get('First Name', '')) and pd.isna(row.get('Last Name', '')):
                        continue
                    
                    # Required field validations
                    first_name = str(row.get('First Name', '')).strip()
                    last_name = str(row.get('Last Name', '')).strip()
                    student_id = str(row.get('Student ID', '')).strip()
                    admission_number = str(row.get('Admission Number', '')).strip()
                    email = str(row.get('Email', '')).strip()
                    
                    if not first_name:
                        raise Exception("First Name is required")
                    if not last_name:
                        raise Exception("Last Name is required")
                    if not student_id:
                        raise Exception("Student ID is required")
                    if not admission_number:
                        raise Exception("Admission Number is required")
                    if not email:
                        raise Exception("Email is required")
                    
                    # Get or create Batch
                    batch_name = str(row.get('Batch Name', '')).strip()
                    if not batch_name:
                        raise Exception("Batch Name is required")
                    batch, batch_created = Batch.objects.get_or_create(name=batch_name)
                    if batch_created:
                        warnings.append(f"Row {index + 2}: Created new batch '{batch_name}'")
                    
                    # Get or create Semester
                    try:
                        semester_num = int(float(row.get('Semester Number', 1)))
                    except:
                        semester_num = 1
                    semester, semester_created = Semester.objects.get_or_create(number=semester_num)
                    if semester_created:
                        warnings.append(f"Row {index + 2}: Created new semester '{semester_num}'")
                    
                    # Get or create Section
                    section_name = str(row.get('Section Name', '')).strip()
                    if not section_name:
                        raise Exception("Section Name is required")
                    section, section_created = Section.objects.get_or_create(
                        name=section_name, 
                        batch=batch
                    )
                    if section_created:
                        warnings.append(f"Row {index + 2}: Created new section '{section_name}' for batch '{batch_name}'")
                    
                    # Get or create Discipline
                    discipline_name = str(row.get('Discipline Name', '')).strip()
                    if not discipline_name:
                        raise Exception("Discipline Name is required")
                    discipline, discipline_created = Discipline.objects.get_or_create(
                        field=discipline_name,
                        defaults={'program': 'General'}
                    )
                    if discipline_created:
                        warnings.append(f"Row {index + 2}: Created new discipline '{discipline_name}'")
                    
                    # Check if student already exists
                    if Student.objects.filter(student_id=student_id).exists():
                        errors.append(f"Row {index + 2}: Student ID '{student_id}' already exists")
                        error_count += 1
                        continue
                    
                    if Student.objects.filter(email=email).exists():
                        errors.append(f"Row {index + 2}: Email '{email}' already exists")
                        error_count += 1
                        continue
                    
                    # Parse gender
                    gender_value = str(row.get('Gender', 'M')).strip().upper()
                    if gender_value in ['M', 'MALE']:
                        gender = 'M'
                    elif gender_value in ['F', 'FEMALE']:
                        gender = 'F'
                    elif gender_value in ['O', 'OTHER']:
                        gender = 'O'
                    else:
                        gender = 'M'  # Default to Male
                    
                    # Parse date of birth
                    dob_value = row.get('Date of Birth')
                    if pd.isna(dob_value):
                        raise Exception("Date of Birth is required")
                    
                    if isinstance(dob_value, str):
                        dob = datetime.strptime(dob_value, '%Y-%m-%d').date()
                    elif isinstance(dob_value, datetime):
                        dob = dob_value.date()
                    else:
                        dob = pd.to_datetime(dob_value).date()
                    
                    # Create Parent (if father name provided)
                    parent = None
                    father_name = str(row.get('Father Name', '')).strip()
                    if father_name:
                        parent = Parent.objects.create(
                            father_name=father_name,
                            mother_name=str(row.get('Mother Name', '')),
                            father_email=str(row.get('Father Email', '')),
                            father_contact=str(row.get('Father Contact', '')),
                            address=str(row.get('Parent Address', ''))
                        )
                    
                    # Get address (make it required)
                    address = str(row.get('Address', '')).strip()
                    if not address:
                        address = "Not provided"
                    
                    # Create Student
                    student = Student(
                        first_name=first_name,
                        last_name=last_name,
                        student_id=student_id,
                        admission_number=admission_number,
                        gender=gender,
                        dob=dob,
                        email=email,
                        contact_number=str(row.get('Contact Number', '')),
                        batch=batch,
                        semester=semester,
                        section=section,
                        discipline=discipline,
                        parent=parent,
                        address=address
                    )
                    student.save()
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    errors.append(f"Row {index + 2}: {str(e)}")
            
            # Show results
            if success_count > 0:
                messages.success(request, f'✅ Successfully imported {success_count} students!')
            
            if warnings:
                for warning in warnings[:5]:
                    messages.warning(request, f'⚠️ {warning}')
            
            if errors:
                for error in errors[:10]:
                    messages.error(request, f'❌ {error}')
                if len(errors) > 10:
                    messages.warning(request, f'And {len(errors) - 10} more errors...')
                    
        except Exception as e:
            messages.error(request, f'Error reading file: {str(e)}')
        
        return redirect('student_list')
    
    return redirect('add_student')


def import_students_excel(request):
    """Alias for bulk_student_import - for template compatibility"""
    return bulk_student_import(request)


def export_students_excel(request):
    """Export all students to Excel"""
    
    students = Student.objects.all().select_related('batch', 'semester', 'section', 'discipline', 'parent')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Students List"
    
    headers = [
        'Student ID', 'First Name', 'Last Name', 'Gender', 'Date of Birth',
        'Email', 'Contact Number', 'Batch', 'Semester', 'Section', 'Discipline',
        'Address', 'Father Name', 'Mother Name', 'Father Contact', 'Father Email'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
    
    for row_num, student in enumerate(students, 2):
        ws.cell(row=row_num, column=1, value=student.student_id)
        ws.cell(row=row_num, column=2, value=student.first_name)
        ws.cell(row=row_num, column=3, value=student.last_name)
        ws.cell(row=row_num, column=4, value=student.get_gender_display())
        ws.cell(row=row_num, column=5, value=student.dob.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=6, value=student.email)
        ws.cell(row=row_num, column=7, value=student.contact_number or '')
        ws.cell(row=row_num, column=8, value=student.batch.name if student.batch else '')
        ws.cell(row=row_num, column=9, value=student.semester.number if student.semester else '')
        ws.cell(row=row_num, column=10, value=student.section.name if student.section else '')
        ws.cell(row=row_num, column=11, value=student.discipline.field if student.discipline else '')
        ws.cell(row=row_num, column=12, value=student.address)
        
        if student.parent:
            ws.cell(row=row_num, column=13, value=student.parent.father_name or '')
            ws.cell(row=row_num, column=14, value=student.parent.mother_name or '')
            ws.cell(row=row_num, column=15, value=student.parent.father_contact or '')
            ws.cell(row=row_num, column=16, value=student.parent.father_email or '')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="students_export.xlsx"'
    wb.save(response)
    return response


# ============= STUDENT MANAGEMENT VIEWS =============

def add_student(request):
    batches = Batch.objects.all()
    semesters = Semester.objects.all()
    sections = Section.objects.all()
    disciplines = Discipline.objects.all()
    
    if request.method == 'POST':
        # Check if this is a bulk import
        if request.FILES.get('excel_file'):
            return bulk_student_import(request)
        
        try:
            # Handle parent data
            parent = None
            if request.POST.get('father_name'):
                parent = Parent.objects.create(
                    father_name=request.POST.get('father_name'),
                    mother_name=request.POST.get('mother_name', ''),
                    father_email=request.POST.get('father_email', ''),
                    father_contact=request.POST.get('father_contact', ''),
                    address=request.POST.get('parent_address', '')
                )
            
            # Create Student
            student = Student(
                first_name=request.POST.get('first_name'),
                last_name=request.POST.get('last_name'),
                student_id=request.POST.get('student_id'),
                admission_number=request.POST.get('admission_number'),
                gender=request.POST.get('gender'),
                dob=request.POST.get('dob'),
                email=request.POST.get('email'),
                contact_number=request.POST.get('contact_number'),
                batch=Batch.objects.get(id=request.POST.get('batch')),
                semester=Semester.objects.get(id=request.POST.get('semester')),
                section=Section.objects.get(id=request.POST.get('section')),
                discipline=Discipline.objects.get(id=request.POST.get('discipline')),
                parent=parent,
                address=request.POST.get('address'),
            )
            
            if 'image' in request.FILES:
                student.image = request.FILES['image']
            
            student.save()
            messages.success(request, 'Student created successfully!')
            return redirect('student_list')
            
        except Exception as e:
            messages.error(request, f'Error creating student: {str(e)}')
    
    return render(request, 'students/add-student.html', {
        'batches': batches,
        'semesters': semesters,
        'sections': sections,
        'disciplines': disciplines,
    })


def student_list(request):
    batches = Batch.objects.all()
    all_sections = Section.objects.all()
    semesters = Semester.objects.all()
    disciplines = Discipline.objects.all()
    total_semesters = semesters.count()

    selected_discipline = None
    selected_batch = None
    selected_section = None
    selected_semester = None

    students = Student.objects.all().order_by('last_name')
    
    # Calculate graduated count (students in semester 8)
    graduated_count = students.filter(semester__number=8).count()

    discipline_id = request.GET.get('discipline')
    if discipline_id:
        selected_discipline = get_object_or_404(Discipline, id=discipline_id)
        students = students.filter(discipline=selected_discipline)

    batch_id = request.GET.get('batch')
    if batch_id:
        selected_batch = get_object_or_404(Batch, id=batch_id)
        students = students.filter(batch=selected_batch)
        sections = all_sections.filter(batch=selected_batch)
    else:
        sections = all_sections

    section_id = request.GET.get('section')
    if section_id:
        selected_section = get_object_or_404(Section, id=section_id)
        students = students.filter(section=selected_section)

    semester_id = request.GET.get('semester')
    if semester_id:
        selected_semester = get_object_or_404(Semester, id=semester_id)
        students = students.filter(semester=selected_semester)

    context = {
        'students': students,
        'batches': batches,
        'sections': sections,
        'semesters': semesters,
        'disciplines': disciplines,
        'total_semesters': total_semesters,
        'graduated_count': graduated_count,
        'selected_discipline': selected_discipline,
        'selected_batch': selected_batch,
        'selected_section': selected_section,
        'selected_semester': selected_semester,
    }
    return render(request, 'students/students.html', context)
def view_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    return render(request, 'students/student-details.html', {'student': student})


def edit_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    batches = Batch.objects.all()
    semesters = Semester.objects.all()
    sections = Section.objects.all()
    disciplines = Discipline.objects.all()
    
    if request.method == 'POST':
        try:
            student.first_name = request.POST.get('first_name')
            student.last_name = request.POST.get('last_name')
            student.student_id = request.POST.get('student_id')
            student.admission_number = request.POST.get('admission_number')
            student.gender = request.POST.get('gender')
            student.dob = request.POST.get('dob')
            student.email = request.POST.get('email')
            student.contact_number = request.POST.get('contact_number')
            student.batch = Batch.objects.get(id=request.POST.get('batch'))
            student.semester = Semester.objects.get(id=request.POST.get('semester'))
            student.section = Section.objects.get(id=request.POST.get('section'))
            student.discipline = Discipline.objects.get(id=request.POST.get('discipline'))
            student.address = request.POST.get('address')
            
            if student.parent:
                student.parent.father_name = request.POST.get('father_name', '')
                student.parent.mother_name = request.POST.get('mother_name', '')
                student.parent.father_email = request.POST.get('father_email', '')
                student.parent.father_contact = request.POST.get('father_contact', '')
                student.parent.address = request.POST.get('parent_address', '')
                student.parent.save()
            elif request.POST.get('father_name'):
                parent = Parent.objects.create(
                    father_name=request.POST.get('father_name'),
                    mother_name=request.POST.get('mother_name', ''),
                    father_email=request.POST.get('father_email', ''),
                    father_contact=request.POST.get('father_contact', ''),
                    address=request.POST.get('parent_address', '')
                )
                student.parent = parent
            
            if 'image' in request.FILES:
                student.image = request.FILES['image']
            
            student.save()
            messages.success(request, 'Student updated successfully!')
            return redirect('student_list')
            
        except Exception as e:
            messages.error(request, f'Error updating student: {str(e)}')
    
    return render(request, 'students/edit-student.html', {
        'student': student,
        'batches': batches,
        'semesters': semesters,
        'sections': sections,
        'disciplines': disciplines,
    })


def delete_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    if request.method == 'POST':
        try:
            if student.parent:
                student.parent.delete()
            student.delete()
            messages.success(request, 'Student deleted successfully!')
        except Exception as e:
            messages.error(request, f'Error deleting student: {str(e)}')
    return redirect('student_list')


def promote_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    
    if request.method == 'POST':
        current_semester = student.semester
        
        if not current_semester:
            messages.error(request, f'Student has no current semester assigned.')
            return redirect('student_list')
        
        # If student is in semester 8, mark as graduated
        if current_semester.number >= 8:
            messages.warning(request, f'🎓 {student.first_name} {student.last_name} has already graduated!')
            return redirect('student_list')
        
        all_semesters = Semester.objects.all().order_by('number')
        next_semester = None
        for semester in all_semesters:
            if semester.number > current_semester.number:
                next_semester = semester
                break
        
        if next_semester:
            student.semester = next_semester
            student.save()
            
            # Check if promoted to semester 8
            if next_semester.number == 8:
                messages.success(request, f'🎓 CONGRATULATIONS! {student.first_name} {student.last_name} is now in final semester (Semester 8)!')
            else:
                messages.success(request, f'✅ {student.first_name} {student.last_name} promoted from Semester {current_semester.number} to Semester {next_semester.number}!')
        else:
            messages.warning(request, f'⚠️ {student.first_name} {student.last_name} has completed all semesters.')
    
    return redirect('student_list')


def promote_all_students(request):
    if request.method == 'POST':
        semester_id = request.POST.get('semester_id')
        current_semester = get_object_or_404(Semester, id=semester_id)
        
        # Check if trying to promote semester 8
        if current_semester.number >= 8:
            messages.warning(request, f'Semester {current_semester.number} is the final semester. Students cannot be promoted further.')
            return redirect('student_list')
        
        all_semesters = Semester.objects.all().order_by('number')
        next_semester = None
        for semester in all_semesters:
            if semester.number > current_semester.number:
                next_semester = semester
                break
        
        if next_semester:
            students_to_promote = Student.objects.filter(semester=current_semester)
            count = students_to_promote.count()
            
            if count > 0:
                students_to_promote.update(semester=next_semester)
                
                if next_semester.number == 9:
                    messages.success(request, f'🎓 Successfully promoted {count} student(s) to FINAL SEMESTER (Semester 8)!')
                else:
                    messages.success(request, f'✅ Successfully promoted {count} student(s) from Semester {current_semester.number} to Semester {next_semester.number}!')
            else:
                messages.info(request, f'No students found in Semester {current_semester.number} to promote.')
        else:
            messages.warning(request, f'Semester {current_semester.number} is the final semester.')
    
    return redirect('student_list')

def promote_all_students(request):
    if request.method == 'POST':
        semester_id = request.POST.get('semester_id')
        current_semester = get_object_or_404(Semester, id=semester_id)
        
        all_semesters = Semester.objects.all().order_by('number')
        next_semester = None
        for semester in all_semesters:
            if semester.number > current_semester.number:
                next_semester = semester
                break
        
        if next_semester:
            students_to_promote = Student.objects.filter(semester=current_semester)
            count = students_to_promote.count()
            
            if count > 0:
                students_to_promote.update(semester=next_semester)
                messages.success(request, f'✅ Successfully promoted {count} student(s) from Semester {current_semester.number} to Semester {next_semester.number}!')
            else:
                messages.info(request, f'No students found in Semester {current_semester.number} to promote.')
        else:
            messages.warning(request, f'Semester {current_semester.number} is the final semester.')
    
    return redirect('student_list')


def get_sections_by_batch(request):
    """AJAX view to get sections for a specific batch"""
    batch_id = request.GET.get('batch_id')
    if batch_id:
        sections = Section.objects.filter(batch_id=batch_id).values('id', 'name')
        return JsonResponse(list(sections), safe=False)
    return JsonResponse([], safe=False)